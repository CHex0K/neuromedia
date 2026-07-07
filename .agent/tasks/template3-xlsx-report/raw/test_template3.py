# -*- coding: utf-8 -*-
"""Unit test for template3_report.build_template3_workbook (no GPU, no media).

The template is formula-driven: only 1_RAW_CORRELATIONS holds raw numbers, while
3_INDEX_SCORES / 4_WINDOW_AGGREGATES / 2_FRAME_MARKUP(A:C) are Excel formulas.
This test verifies the generator fills the raw sheet per segment, PRESERVES all
formulas, and clears phantom frame rows beyond the real segment count (so empty
formula rows cannot poison the STDEV/AVERAGE ranges).
"""

import sys
import tempfile
from pathlib import Path

import openpyxl
import pandas as pd

REPO_ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(REPO_ROOT))

import template3_report as t3  # noqa: E402

TEMPLATE = REPO_ROOT / "Template_3_with_frames.xlsx"


def _make_inputs(tmp: Path):
    term_rows = []
    feats = {"attention": "attention", "visual attention": "visual_attention", "reward": "reward"}
    for ti in range(3):  # 3 segments -> F01..F03 (< 20, trim path)
        for raw_feat in feats:
            grp = "attention" if raw_feat != "reward" else "reward"
            term_rows.append({
                "map_id": "m", "map_path": "m.npy", "time_index": ti, "group": grp,
                "alias": raw_feat, "feature": raw_feat, "match_type": "reference",
                "r": 0.1 * ti + (0.5 if raw_feat == "reward" else 0.2),
            })
    pd.DataFrame(term_rows).to_csv(tmp / "decoded_terms.csv", index=False)
    # mimic the REAL tribe_segments.tsv: offset column empty, real time in `start`
    pd.DataFrame([
        {"index": 0, "offset": None, "duration": 1.0, "start": 0.0},
        {"index": 1, "offset": None, "duration": 1.0, "start": 1.0},
        {"index": 2, "offset": None, "duration": 1.0, "start": 2.0},
    ]).to_csv(tmp / "tribe_segments.tsv", sep="\t", index=False)
    # transcript words aligned to those start windows
    pd.DataFrame([
        {"start": 0.2, "text": "hello"},
        {"start": 1.2, "text": "world"},
        {"start": 2.2, "text": "foo"},
    ]).to_csv(tmp / "words.tsv", sep="\t", index=False)


def main() -> None:
    with tempfile.TemporaryDirectory() as d:
        tmp = Path(d)
        _make_inputs(tmp)
        out = tmp / "Template_3_out.xlsx"
        t3.build_template3_workbook(
            template_path=TEMPLATE,
            decoded_terms_csv=tmp / "decoded_terms.csv",
            segments_tsv=tmp / "tribe_segments.tsv",
            words_tsv=tmp / "words.tsv",
            output_xlsx=out,
            frame_png_provider=None,
        )
        wb = openpyxl.load_workbook(out, data_only=False)  # keep formulas

        assert {"1_RAW_CORRELATIONS", "2_FRAME_MARKUP", "3_INDEX_SCORES",
                "4_WINDOW_AGGREGATES", "5_WEIGHTS_REF", "README"}.issubset(set(wb.sheetnames))

        # --- 1_RAW_CORRELATIONS: raw numbers written; phantom rows cleared ---
        s1 = wb["1_RAW_CORRELATIONS"]
        assert s1["A4"].value == "F01" and s1["A5"].value == "F02" and s1["A6"].value == "F03"
        assert s1["A7"].value is None, "rows beyond segment count must be cleared"
        # timing must come from the populated `start` column (offset was empty), and
        # scene_description must be aligned from the words TSV -> proves the fix
        assert s1["B4"].value == 0.0 and s1["B5"].value == 1.0 and s1["B6"].value == 2.0, "second from start"
        assert s1["C4"].value == "hello" and s1["C5"].value == "world" and s1["C6"].value == "foo", \
            f"scene_description not aligned: {[s1['C4'].value, s1['C5'].value, s1['C6'].value]}"
        assert s1.cell(3, 6).value == "visual_attention" and s1.cell(3, 30).value == "reward"
        assert abs(float(s1.cell(6, 6).value) - 0.4) < 1e-9   # visual_attention @ ti=2
        assert abs(float(s1.cell(4, 30).value) - 0.5) < 1e-9  # reward @ ti=0

        # --- 3_INDEX_SCORES: formulas PRESERVED for real rows, phantom rows cleared ---
        s3 = wb["3_INDEX_SCORES"]
        assert isinstance(s3["D4"].value, str) and s3["D4"].value.startswith("=AVERAGE"), s3["D4"].value
        assert isinstance(s3["L4"].value, str) and s3["L4"].value.startswith("=IFERROR"), s3["L4"].value
        assert isinstance(s3["A4"].value, str) and s3["A4"].value.startswith("="), "frame_id link preserved"
        assert s3["D7"].value is None, "phantom index formula rows (>3 frames) must be cleared"
        # template ships a red example fill on C6 (F03 tags) -> must be stripped
        assert s3["C6"].fill.patternType != "solid", ("example red fill leaked", s3["C6"].fill.patternType)

        # --- 2_FRAME_MARKUP: linked headers preserved, flag skeleton present, phantom cleared ---
        s2 = wb["2_FRAME_MARKUP"]
        assert isinstance(s2["A4"].value, str) and s2["A4"].value.startswith("="), "markup frame_id link preserved"
        assert s2["E4"].value == 0  # HOOK skeleton reset
        # template ships example LOGO=1 on F03/F04 (rows 6/7 area) -> must be reset to 0
        assert s2["J6"].value == 0, ("template example markup leaked into output", s2["J6"].value)
        assert s2["A7"].value is None, "phantom markup rows cleared"

        # --- 4_WINDOW_AGGREGATES: untouched formulas ---
        s4 = wb["4_WINDOW_AGGREGATES"]
        assert isinstance(s4["C4"].value, str) and s4["C4"].value.startswith("=AVERAGE"), s4["C4"].value

        print("PASS: raw sheet filled; index/window formulas preserved; phantom rows cleared")


if __name__ == "__main__":
    main()
    print("TEMPLATE3 TEST PASSED")
