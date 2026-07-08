# -*- coding: utf-8 -*-
"""Build a Template_3_with_frames.xlsx report from TRIBE pipeline artifacts.

The shipped ``Template_3_with_frames.xlsx`` is FORMULA-DRIVEN: only
``1_RAW_CORRELATIONS`` holds raw numbers (per-segment x per-term Pearson r);
``2_FRAME_MARKUP`` (frame_id/second/scene), ``3_INDEX_SCORES`` and
``4_WINDOW_AGGREGATES`` are Excel formulas that reference sheet 1 and the
``5_WEIGHTS_REF`` weights. So this generator fills ONLY the raw-correlations
sheet (plus the manual-markup skeleton and embedded frames) and leaves every
formula untouched — Excel recomputes the indices, z-scores and window aggregates
when the file is opened.

Frame rows in the skeleton are F01..F20 (rows 4..23). We fit them to the actual
number of TRIBE segments: extra rows are cleared (so empty formula rows cannot
turn into #DIV/0! and poison the STDEV/AVERAGE ranges), and if there are more
than 20 segments the formula rows are extended.

Data sources (all produced by the existing pipeline):
  - decoded_terms.csv    : per-segment (time_index) x per-term (feature) Pearson r
  - tribe_segments.tsv   : segment timing (index/offset/duration)
  - words TSV (optional) : transcript words for scene_description alignment
  - frame_png_provider   : optional callable(offset, duration) -> PNG bytes | None
"""

from __future__ import annotations

import io
import logging
from pathlib import Path
from typing import Callable, Optional

import pandas as pd

LOGGER = logging.getLogger("template3_report")

DATA_START_ROW = 4          # frame rows start at row 4 (headers on rows 2-3)
TEMPLATE_LAST_FRAME_ROW = 23  # skeleton has F01..F20 in rows 4..23
FRAME_COL = 4               # column D = frame_image
FIRST_TERM_COL = 5          # raw-correlation term columns start at E
MARKUP_FLAG_COLS = range(5, 11)  # HOOK/BRAND/PRODUCT/OFFER/CTA/LOGO
FINALE_WINDOW_SECONDS = 3.0
FINALE_AGGREGATE_ROW = 8
FIRST_AGGREGATE_VALUE_COL = 3  # column C on 4_WINDOW_AGGREGATES
SCENE_DESCRIPTION_COL = 3
SCENE_SELECTED_TEXT_COLUMNS = ["text", "word", "token", "corrected", "corrected_text"]
SCENE_ORIGINAL_TEXT_COLUMNS = ["original_text", "source_text", "raw_text"]
RUSSIAN_LANGUAGE_VALUES = {"ru", "rus", "russian", "русский"}
RED_FILL_RGBS = {"FFFF0000", "00FF0000", "FF0000"}

FramePngProvider = Callable[[Optional[float], Optional[float]], Optional[bytes]]


def _norm_term(term: object) -> str:
    return str(term or "").strip().lower().replace(" ", "_").replace("-", "_")


def _find_column(frame: pd.DataFrame, candidates: list[str]) -> Optional[str]:
    normalized = {str(c).strip().lower(): c for c in frame.columns}
    for cand in candidates:
        if cand in normalized:
            return normalized[cand]
    return None


def _first_populated_column(frame: pd.DataFrame, candidates: list[str]) -> Optional[str]:
    """Pick the first candidate column that exists AND has data.

    TRIBE's tribe_segments.tsv leaves ``offset`` empty and stores the real
    per-segment time in ``start``; a plain name match would pick the empty
    ``offset`` column, so prefer a candidate that actually has values.
    """

    normalized = {str(c).strip().lower(): c for c in frame.columns}
    for cand in candidates:
        col = normalized.get(cand)
        if col is not None and frame[col].notna().any():
            return col
    for cand in candidates:  # fall back to any existing (even if empty)
        if cand in normalized:
            return normalized[cand]
    return None


def _has_text_values(frame: pd.DataFrame, column: str) -> bool:
    values = frame[column].dropna().astype(str).str.strip()
    return bool(values.ne("").any())


def _target_language_is_russian(frame: pd.DataFrame) -> bool:
    column = _find_column(frame, ["target_language"])
    if column is None:
        return False
    values = {
        str(value).strip().lower().replace("_", "-")
        for value in frame[column].dropna().unique()
    }
    return bool(values & RUSSIAN_LANGUAGE_VALUES)


def _scene_description_text_column(frame: pd.DataFrame) -> Optional[str]:
    """Pick display text for Excel without changing TRIBE scoring text."""

    selected_text = _first_populated_column(frame, SCENE_SELECTED_TEXT_COLUMNS)
    original_text = _first_populated_column(frame, SCENE_ORIGINAL_TEXT_COLUMNS)
    if _target_language_is_russian(frame) and selected_text and _has_text_values(frame, selected_text):
        return selected_text
    if original_text and _has_text_values(frame, original_text):
        return original_text
    if selected_text and _has_text_values(frame, selected_text):
        return selected_text
    return original_text or selected_text


def _read_header_terms(ws) -> list[tuple[int, str]]:
    """Return (column_index, normalized_term) for the term columns of sheet 1."""

    terms: list[tuple[int, str]] = []
    for col in range(FIRST_TERM_COL, ws.max_column + 1):
        name = ws.cell(3, col).value
        if name:
            terms.append((col, _norm_term(name)))
    return terms


def _segment_rows(decoded_terms_csv, segments_tsv, words_tsv):
    """Assemble per-segment (frame) structures from the pipeline artifacts."""

    terms = pd.read_csv(decoded_terms_csv, encoding="utf-8")
    terms["_feat"] = terms["feature"].map(_norm_term)
    r_by_ti_feat = (
        terms.drop_duplicates(subset=["time_index", "_feat"])
        .set_index(["time_index", "_feat"])["r"]
        .to_dict()
    )
    time_indices = sorted(int(t) for t in pd.unique(terms["time_index"]))

    offsets: dict[int, Optional[float]] = {}
    durations: dict[int, Optional[float]] = {}
    if segments_tsv and Path(segments_tsv).is_file():
        seg = pd.read_csv(segments_tsv, sep="\t", encoding="utf-8")
        off_col = _first_populated_column(seg, ["offset", "start", "timeline"])
        dur_col = _first_populated_column(seg, ["duration"])
        idx_col = _find_column(seg, ["index"])
        for i, row in seg.iterrows():
            ti = int(row[idx_col]) if idx_col is not None and pd.notna(row[idx_col]) else i
            offsets[ti] = float(row[off_col]) if off_col and pd.notna(row[off_col]) else None
            durations[ti] = float(row[dur_col]) if dur_col and pd.notna(row[dur_col]) else None

    words = None
    if words_tsv and Path(words_tsv).is_file():
        wdf = pd.read_csv(words_tsv, sep="\t", encoding="utf-8")
        ws_col = _find_column(wdf, ["start", "word_start", "begin"])
        wt_col = _scene_description_text_column(wdf)
        if ws_col and wt_col:
            words = wdf[[ws_col, wt_col]].rename(columns={ws_col: "start", wt_col: "text"})

    def scene_desc(offset, duration):
        if words is None or offset is None:
            return ""
        end = offset + (duration or 0.0)
        sel = words[(words["start"] >= offset) & (words["start"] < max(end, offset + 0.001))]
        return " ".join(str(t) for t in sel["text"].tolist()).strip()

    rows = []
    for pos, ti in enumerate(time_indices):
        offset = offsets.get(ti)
        rows.append({
            "time_index": ti,
            "frame_id": f"F{pos + 1:02d}",
            "second": round(offset, 2) if offset is not None else pos + 1,
            "offset": offset,
            "duration": durations.get(ti),
            "scene_description": scene_desc(offset, durations.get(ti)),
            "r": r_by_ti_feat,
        })
    return rows


def _embed_frame(ws, row: int, png_bytes: Optional[bytes]) -> None:
    if not png_bytes:
        return
    try:
        from openpyxl.drawing.image import Image as XLImage

        image = XLImage(io.BytesIO(png_bytes))
        image.width, image.height = 140, 80
        ws.row_dimensions[row].height = 64
        ws.add_image(image, f"D{row}")
    except Exception as exc:  # pragma: no cover - PIL/openpyxl optional at runtime
        LOGGER.warning("Could not embed frame at row %d: %s", row, exc)


def _clear_rows(ws, start_row: int, end_row: int) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).value = None


def _drop_images(ws) -> None:
    if hasattr(ws, "_images"):
        ws._images = []


def _reset_data_fills(ws, start_row: int, end_row: int) -> None:
    """Clear solid cell fills left over from the template example on data rows.

    The shipped template highlights an example cell red (3_INDEX_SCORES!C6, the
    F03 tags cell). Structural reds live in header rows / other sheets and are not
    touched here.
    """

    from openpyxl.styles import PatternFill

    blank = PatternFill(fill_type=None)
    for row in range(start_row, end_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if cell.fill is not None and cell.fill.patternType == "solid":
                cell.fill = blank


def _is_red_fill(fill) -> bool:
    if fill is None or fill.patternType != "solid":
        return False
    rgb = str(getattr(fill.fgColor, "rgb", "") or "").upper()
    return rgb in RED_FILL_RGBS


def _remove_red_fills(wb) -> None:
    """Remove red highlights from generated workbooks only."""

    from openpyxl.styles import PatternFill

    blank = PatternFill(fill_type=None)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if _is_red_fill(cell.fill):
                    cell.fill = blank


def _extend_formula_rows(ws, template_row: int, first_new_row: int, last_new_row: int) -> None:
    """Copy the template's formula/static row down to new rows, shifting refs."""

    from openpyxl.formula.translate import Translator

    for new_row in range(first_new_row, last_new_row + 1):
        for col in range(1, ws.max_column + 1):
            src = ws.cell(template_row, col)
            dst = ws.cell(new_row, col)
            if src.data_type == "f" and isinstance(src.value, str):
                dst.value = Translator(src.value, origin=f"{src.column_letter}{template_row}").translate_formula(
                    f"{src.column_letter}{new_row}"
                )
            elif col >= FIRST_TERM_COL or col <= 3:
                dst.value = src.value  # carry static markup flags / labels


def _widen_bottom_anchor(ws, new_last_row: int) -> None:
    """Rewrite fixed '23' bottom-row anchors in formulas to the real last row."""

    old = str(TEMPLATE_LAST_FRAME_ROW)
    new = str(new_last_row)
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == "f" and isinstance(cell.value, str) and old in cell.value:
                # only touch range anchors like ':D$23' / ':D23' / 'D4:D23'
                cell.value = cell.value.replace("$" + old, "$" + new).replace(old + ")", new + ")")


def _positive_float(value: object) -> Optional[float]:
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if result > 0 else None


def _numeric_float(value: object) -> Optional[float]:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _finale_window_rows(rows: list[dict]) -> tuple[int, int]:
    """Return Excel row bounds for segments overlapping the final 3 seconds."""

    if not rows:
        return DATA_START_ROW, DATA_START_ROW

    fallback_start = DATA_START_ROW + max(len(rows) - 3, 0)
    fallback_end = DATA_START_ROW + len(rows) - 1

    starts = [_numeric_float(row.get("offset")) for row in rows]
    if not any(start is not None for start in starts):
        return fallback_start, fallback_end

    timed: list[tuple[int, float, float]] = []
    for index, row in enumerate(rows):
        start = starts[index]
        if start is None:
            continue

        duration = _positive_float(row.get("duration"))
        next_start = next(
            (
                candidate
                for candidate in starts[index + 1 :]
                if candidate is not None and candidate > start
            ),
            None,
        )
        if duration is not None:
            end = start + duration
        elif next_start is not None:
            end = next_start
        else:
            end = start + 1.0
        if end <= start:
            end = start + 1.0
        timed.append((DATA_START_ROW + index, start, end))

    if not timed:
        return fallback_start, fallback_end

    video_end = max(end for _, _, end in timed)
    window_start = max(0.0, video_end - FINALE_WINDOW_SECONDS)
    selected = [
        excel_row
        for excel_row, start, end in timed
        if end > window_start and start < video_end
    ]
    if not selected:
        selected = [timed[-1][0]]
    return min(selected), max(selected)


def _rewrite_finale_window(ws, rows: list[dict]) -> None:
    """Rewrite Finale aggregate formulas to the actual final 3-second window."""

    from openpyxl.utils import get_column_letter

    first_row, last_row = _finale_window_rows(rows)
    seconds_label = int(FINALE_WINDOW_SECONDS)
    ws.cell(FINALE_AGGREGATE_ROW, 1).value = f"Finale last {seconds_label}s"
    ws.cell(FINALE_AGGREGATE_ROW, 2).value = "Среднее по кадрам за последние 3 секунды"
    for col in range(FIRST_AGGREGATE_VALUE_COL, ws.max_column + 1):
        source_col = get_column_letter(col + 1)  # aggregate C maps to index-score D
        ws.cell(FINALE_AGGREGATE_ROW, col).value = (
            f"=AVERAGE('3_INDEX_SCORES'!{source_col}{first_row}:{source_col}{last_row})"
        )


def _rewrite_markup_scene_formulas(ws, last_data_row: int) -> None:
    """Keep blank scene descriptions blank instead of showing Excel's 0."""

    if last_data_row < DATA_START_ROW:
        return
    for row in range(DATA_START_ROW, last_data_row + 1):
        raw_ref = f"'1_RAW_CORRELATIONS'!C{row}"
        ws.cell(row, SCENE_DESCRIPTION_COL).value = f'=IF({raw_ref}="","",{raw_ref})'


def build_template3_workbook(
    *,
    template_path: Path,
    decoded_terms_csv: Path,
    output_xlsx: Path,
    marketing_scores_csv: Optional[Path] = None,  # kept for call compatibility; unused (formulas compute it)
    segments_tsv: Optional[Path] = None,
    words_tsv: Optional[Path] = None,
    frame_png_provider: Optional[FramePngProvider] = None,
) -> Path:
    """Fill only the raw-correlations sheet; leave all index/window formulas intact."""

    import openpyxl

    wb = openpyxl.load_workbook(template_path)  # keep formulas (data_only=False)
    rows = _segment_rows(decoded_terms_csv, segments_tsv, words_tsv)
    n = len(rows)
    last_data_row = DATA_START_ROW + n - 1

    ws_raw = wb["1_RAW_CORRELATIONS"]
    ws_mark = wb["2_FRAME_MARKUP"]
    ws_idx = wb["3_INDEX_SCORES"]
    ws_agg = wb["4_WINDOW_AGGREGATES"]

    term_cols = _read_header_terms(ws_raw)

    # 1) reset raw sheet body and (re)write one row per segment
    _drop_images(ws_raw)
    _drop_images(ws_mark)
    _clear_rows(ws_raw, DATA_START_ROW, max(ws_raw.max_row, TEMPLATE_LAST_FRAME_ROW))
    for i, seg in enumerate(rows):
        r = DATA_START_ROW + i
        ws_raw.cell(r, 1).value = seg["frame_id"]
        ws_raw.cell(r, 2).value = seg["second"]
        ws_raw.cell(r, 3).value = seg["scene_description"]
        for col, term in term_cols:
            val = seg["r"].get((seg["time_index"], term))
            ws_raw.cell(r, col).value = float(val) if val is not None else None
        # reset the manual markup flags to a clean 0 skeleton (the shipped template
        # ships example values like LOGO=1 on some frames; those must not leak in)
        for col in MARKUP_FLAG_COLS:
            ws_mark.cell(r, col).value = 0
        if frame_png_provider is not None:
            png = frame_png_provider(seg["offset"], seg["duration"])
            _embed_frame(ws_raw, r, png)
            _embed_frame(ws_mark, r, png)

    # 1b) strip leftover example cell fills (e.g. the red F03 tags cell) from data rows
    fill_end = max(last_data_row, TEMPLATE_LAST_FRAME_ROW)
    for ws in (ws_raw, ws_mark, ws_idx):
        _reset_data_fills(ws, DATA_START_ROW, fill_end)

    # 2) fit the formula sheets (2_FRAME_MARKUP, 3_INDEX_SCORES) to exactly n rows
    if n < (TEMPLATE_LAST_FRAME_ROW - DATA_START_ROW + 1):
        # clear phantom frame rows so empty formulas cannot poison STDEV/AVERAGE
        _clear_rows(ws_mark, last_data_row + 1, TEMPLATE_LAST_FRAME_ROW)
        _clear_rows(ws_idx, last_data_row + 1, TEMPLATE_LAST_FRAME_ROW)
    elif n > (TEMPLATE_LAST_FRAME_ROW - DATA_START_ROW + 1):
        # extend formula rows and widen the fixed 23-row aggregation ranges
        for ws in (ws_mark, ws_idx):
            _extend_formula_rows(ws, TEMPLATE_LAST_FRAME_ROW, TEMPLATE_LAST_FRAME_ROW + 1, last_data_row)
        _widen_bottom_anchor(ws_idx, last_data_row)
        _widen_bottom_anchor(ws_agg, last_data_row)

    _rewrite_markup_scene_formulas(ws_mark, last_data_row)
    _rewrite_finale_window(ws_agg, rows)

    # force Excel/LibreOffice to recompute every formula when the file is opened
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:  # pragma: no cover - older openpyxl
        pass

    _remove_red_fills(wb)

    output_xlsx = Path(output_xlsx)
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)
    LOGGER.info("Wrote Template_3 workbook with %d segment rows: %s", n, output_xlsx)
    return output_xlsx
